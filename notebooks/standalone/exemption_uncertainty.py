# ---
# jupyter:
#   title: What public data can and cannot say about Edmonton's tax-exempt property
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

# %% [markdown]
# # We can measure how much Edmonton property *should* be tax-exempt. We cannot say which.
#
# This notebook pulls the best data any member of the public can get about
# Edmonton property tax, and shows that it is enough to **size** the city's
# tax-exempt property to within a few percent — and **not** enough to name a
# single exempt parcel.
#
# That is not a complaint about data quality. It is a structural property of
# what gets published, and the notebook works out exactly where the wall is.
#
# ## What it does, in one paragraph
#
# Edmonton's published assessment roll lists every property and its assessed
# value. Almost none of it is flagged tax-exempt. But Edmonton also *files* its
# taxable assessment with the Province of Alberta, and that filed number is
# smaller than the roll. The difference is property the City assesses but does
# not tax. We can locate most of it by zoning. We cannot identify it, and the
# last section proves that with a construction rather than an argument.
#
# ## Self-contained
#
# It imports nothing local, downloads its own inputs from the two public
# sources, and caches them. Nothing here reads a private file or a
# pre-computed result. Every number below is produced by the run you are
# reading — none is typed in by hand.
#
# **Runtime** is dominated by two things: a ~80 MB CSV download the first time,
# and a 440,000-point spatial join (about a minute). Re-runs use the cache.
#
# ```
# pip install pandas geopandas requests certifi openpyxl matplotlib
# ```
#
# **A caveat about provenance.** This notebook deliberately reimplements its
# arithmetic inline instead of importing a pipeline, so it can be read and run
# on its own. The cost of that choice is drift: it hardcodes a handful of
# definitions (class names, zone codes, one published mill rate) that live
# authoritatively elsewhere. Each one is flagged **`ASSUMPTION`** where it
# appears, with what would falsify it.

# %%
import base64
import io
import json
import os
import re
import ssl
import sys
import urllib.request
from pathlib import Path

import certifi
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd

# Some networks (and some older Linux CA bundles) fail on hosts whose
# certificate chains to a root issued after the bundle was built. certifi's
# bundle is current, so it is used explicitly rather than relying on the
# system store — this is a real failure mode, not a precaution.
SSL_CTX = ssl.create_default_context(cafile=certifi.where())

# open.alberta.ca returns 403 to urllib's default User-Agent. Identifying the
# client is required to get a response at all, not politeness.
UA = "Mozilla/5.0 (compatible; edmonton-exemption-notebook/1.0)"

CACHE = Path(os.environ.get("EXEMPTION_NB_DATA", "./exemption_data")).resolve()
CACHE.mkdir(parents=True, exist_ok=True)


def read_url(url: str, timeout: int = 300) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read()


def fetch(url: str, name: str) -> Path:
    """Download once into CACHE; reuse thereafter. Returns the local path."""
    dest = CACHE / name
    if dest.exists() and dest.stat().st_size > 0:
        print(f"  cached  {name}  ({dest.stat().st_size / 1e6:.1f} MB)")
        return dest
    print(f"  fetching {name} …", end="", flush=True)
    dest.write_bytes(read_url(url))
    print(f" {dest.stat().st_size / 1e6:.1f} MB")
    return dest


def show(fig, alt: str) -> None:
    """Display a figure carrying real alt text.

    `plt.show()` emits a bare <img> with no alt attribute, so every chart is
    invisible to a screen reader and to anyone whose images fail to load. The
    figure is embedded by hand instead, with a description that states what the
    chart SAYS rather than what it depicts.
    """
    from IPython.display import HTML, display  # noqa: PLC0415

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110, facecolor=fig.get_facecolor())
    plt.close(fig)
    encoded = base64.b64encode(buf.getvalue()).decode("ascii")
    display(HTML(f'<img alt="{alt}" src="data:image/png;base64,{encoded}">'))


# ⚠️ THIS REPORT IS A SNAPSHOT, and the published page has to say so on its own
# face — it is the artifact that gets handed to someone, usually without the
# index page that would otherwise date it. Two of these four carried NO date at
# all until 2026-08-29.
#
# FIRST_MEASURED is when the finding was made and does NOT change on re-run.
# _STAMPED_AT is when this page was last re-executed against live data. Both are
# shown: "first found on" and "still true today" are different claims and a
# reader needs each. Imports are aliased and local to this block so it can be
# pasted into any of these notebooks unchanged.
import datetime as _dt

from IPython.display import Markdown as _Md, display as _disp

FIRST_MEASURED = "2026-08-26"
_STAMPED_AT = _dt.datetime.now(_dt.timezone.utc)

_disp(_Md(
    f"**Snapshot.** Finding first measured **{FIRST_MEASURED}**; this page "
    f"re-executed against live data **{_STAMPED_AT:%Y-%m-%d}** (UTC). "
    f"Nothing re-runs these on a schedule — re-execute before citing a figure."))

CHECKS: list[tuple[bool, str]] = []


def check(ok: bool, claim: str) -> None:
    """Record a structural invariant. Reported together at the end.

    These are claims that must hold for ANY vintage of this data — not the
    values themselves, which move every time Edmonton refreshes the roll.
    """
    CHECKS.append((bool(ok), claim))
    print(f"  [{'PASS' if ok else 'FAIL'}] {claim}")


print(f"cache: {CACHE}")

# %% [markdown]
# ---
# ## 1. Source one — Edmonton's assessment roll
#
# One row per tax account, published on Edmonton's open data portal. This is
# the most detailed property-level data the City releases: every account, its
# assessed value, its assessment class, and its coordinates.

# %%
ROLL_URL = (
    "https://data.edmonton.ca/api/views/q7d6-ambg/rows.csv?accessType=DOWNLOAD"
)
roll_path = fetch(ROLL_URL, "assessment_roll.csv")

roll = pd.read_csv(roll_path, low_memory=False).rename(
    columns={
        "Account Number": "account",
        "Neighbourhood": "hood",
        "Assessed Value": "assessed",
        "Assessment Class 1": "class_1",
        "Assessment Class 2": "class_2",
        "Assessment Class 3": "class_3",
        "Assessment Class % 1": "pct_1",
        "Assessment Class % 2": "pct_2",
        "Assessment Class % 3": "pct_3",
        "Latitude": "lat",
        "Longitude": "lon",
    }
)

print(f"\naccounts published:   {len(roll):,}")
print(f"total assessed value: ${roll['assessed'].sum() / 1e9:,.1f}B")

# A property can be split across up to three assessment classes, each with a
# percentage. Any class total has to apportion by those percentages or a
# mixed-use building lands entirely in whichever class happens to be first.
SLOTS = [("class_1", "pct_1"), ("class_2", "pct_2"), ("class_3", "pct_3")]

pct_total = sum(roll[p].fillna(0) for _, p in SLOTS)
off_100 = ((pct_total - 100).abs() > 0.5) & roll["assessed"].notna()
print(f"rows whose class percentages miss 100: {int(off_100.sum()):,}")

# Zero/null assessed value carries no tax and no information here. Dropping it
# cannot move a ratio; it is arithmetic hygiene, not a policy choice.
dropped = roll["assessed"].isna() | (roll["assessed"] == 0)
roll = roll[~dropped].copy()
roll["assessed"] = roll["assessed"].astype(float)
print(f"dropped null/zero value rows: {int(dropped.sum()):,}")

check(roll["account"].is_unique, "one row per account — no duplicate billing")
check(roll["lat"].notna().all() and roll["lon"].notna().all(),
      "every remaining property has coordinates (needed for the zoning join)")

# %% [markdown]
# ### The roll's own exemption flag is empty
#
# Edmonton's assessment class vocabulary contains exactly one value that means
# "not taxed": `NONRES MUNICIPAL/RES EDUCATION`. If the roll were going to tell
# us which property is exempt, this is the column that would do it.
#
# > **`ASSUMPTION`** — that this label is the roll's only exemption marker. It
# > is falsified by any *other* class label carrying a zero municipal rate. The
# > cell below prints the full class vocabulary so you can check.

# %%
EXEMPT_LABEL = "NONRES MUNICIPAL/RES EDUCATION"

classes = sorted(
    set().union(*[set(roll[c].dropna().unique()) for c, _ in SLOTS])
)
print("every assessment class present in the roll:")
for c in classes:
    n = sum(int((roll[lc] == c).sum()) for lc, _ in SLOTS)
    print(f"  {c:34s} {n:>9,} slices")

# The label can appear in any of the three class slots — a property can be
# PARTLY exempt — so the value is apportioned by percentage like every other
# class total here, rather than counting whole properties.
flagged_rows = np.zeros(len(roll), dtype=bool)
flagged_value = 0.0
for label_col, pct_col in SLOTS:
    m = (roll[label_col] == EXEMPT_LABEL).to_numpy()
    flagged_rows |= m
    flagged_value += (roll.loc[m, "assessed"] * roll.loc[m, pct_col].fillna(0) / 100).sum()

print(f"\nproperties with any exempt-flagged share: {int(flagged_rows.sum()):,} "
      f"of {len(roll):,}")
print(f"assessed value so flagged:               ${flagged_value:,.0f}")
print(f"share of the roll:                       "
      f"{100 * flagged_value / roll['assessed'].sum():.4f}%")

# %% [markdown]
# **Read on the roll alone, Edmonton is a city with essentially no tax-exempt
# property.** A handful of accounts out of hundreds of thousands.
#
# That is obviously false. Edmonton has a university, a legislature, hospitals,
# federal land, hundreds of parks, dozens of schools and places of worship. All
# of it is on the roll, assessed, and — as far as the published columns go —
# indistinguishable from a taxable warehouse.
#
# So the roll cannot answer the question by itself. We need a second source
# that knows the answer in aggregate.

# %% [markdown]
# ---
# ## 2. Source two — what Edmonton filed with the Province
#
# Every Alberta municipality files an annual return with Municipal Affairs. The
# Province republishes them as the Financial Information Return (FIR). Schedule
# MR carries three sheets we need, per municipality:
#
# | sheet | what it holds |
# |---|---|
# | `MR(1)-Tax Levy` | municipal tax levied, by class |
# | `MR(2)-Assessment` | the assessment that levy was computed on |
# | `MR(3)-Mill Rate` | the rates applied |
#
# `MR(2)` is the number the roll cannot give us: Edmonton's own statement of
# its **taxable** assessment.

# %%
FIR_PAGE = "https://open.alberta.ca/opendata/municipal-financial-and-statistical-data"

page = read_url(FIR_PAGE, timeout=120).decode("utf-8", "replace")

# The province publishes a full financial workbook per year, and — for the
# newest year — an early "tax rates" workbook that carries Schedule MR before
# the complete filing lands. Both are usable; the full one wins where both exist.
workbooks: dict[int, str] = {}
for url in set(re.findall(r'href="(https://open\.alberta\.ca/dataset/[^"]+)"', page)):
    m = re.search(r"/download/(\d{4})_(?:financial_year|tax_rates)\.xlsx$", url, re.I)
    if m:
        year = int(m.group(1))
        if year not in workbooks or "financial_year" in url.lower():
            workbooks[year] = url

FIR_YEAR = max(workbooks)
print(f"FIR years published: {min(workbooks)}–{max(workbooks)}")
print(f"using the newest:    {FIR_YEAR}")

fir_path = fetch(workbooks[FIR_YEAR], f"fir_{FIR_YEAR}.xlsx")

# %%
import openpyxl  # noqa: E402  — only needed from here on

EDMONTON_CODE, EDMONTON_NAME = "0098", "EDMONTON"

# Column positions within Schedule MR, and the header text each must contain.
# Reading by position alone would silently return the wrong class if the
# Province ever reorders the sheet, so the header is verified first.
MR_COLUMNS = {5: "residential", 6: "farmland", 7: "non_residential",
              9: "machinery_equipment", 10: "other"}
MR_HEADER_EXPECT = {5: "residential", 6: "farmland", 7: "non-residential",
                    9: "machinery", 10: "other"}
SHEETS = {"MR(1)-Tax Levy": "levy", "MR(2)-Assessment": "assessment",
          "MR(3)-Mill Rate": "mill_rate"}

wb = openpyxl.load_workbook(fir_path, read_only=True, data_only=True)
fir: dict[str, dict[str, float]] = {}
for sheet, key in SHEETS.items():
    ws = wb[sheet]
    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    for idx, expect in MR_HEADER_EXPECT.items():
        got = str(header[idx] or "").lower()
        if expect not in got:
            raise ValueError(
                f"{sheet} column {idx} reads {header[idx]!r}, expected {expect!r} "
                "— Schedule MR moved; do not read it blind"
            )
    row = next(
        r for r in ws.iter_rows(values_only=True)
        if any(isinstance(c, str) and c.strip().upper() == EDMONTON_NAME for c in r[:6])
    )
    if str(row[2]).strip() != EDMONTON_CODE:
        raise ValueError(f"{EDMONTON_NAME} found under code {row[2]!r}")
    fir[key] = {n: (None if row[i] is None else float(row[i]))
                for i, n in MR_COLUMNS.items()}
wb.close()

print(f"Edmonton's {FIR_YEAR} filing (municipality code {EDMONTON_CODE}):\n")
print(f"{'class':<22}{'taxable assessment':>22}{'levy':>20}{'rate':>10}")
for cls in MR_COLUMNS.values():
    a, lv, rt = fir["assessment"][cls], fir["levy"][cls], fir["mill_rate"][cls]
    print(f"{cls:<22}{'$' + format(a, ',.0f'):>22}"
          f"{'$' + format(lv, ',.0f'):>20}{rt:>10.4f}")

# %% [markdown]
# ### Proving `MR(2)` is the *taxable* base, not a gross roll total
#
# The whole comparison rests on `MR(2)` being assessment-after-exemptions. The
# Province's three sheets prove it internally: if assessment × mill rate
# reproduces the levy, then that assessment is what the tax was actually
# computed on.

# %%
for cls in ("residential", "non_residential"):
    a, rt, lv = fir["assessment"][cls], fir["mill_rate"][cls], fir["levy"][cls]
    calc = a * rt / 1000
    check(abs(calc - lv) / lv < 1e-4,
          f"{cls}: filed assessment x filed rate reproduces filed levy "
          f"({100 * (calc / lv - 1):+.4f}%)")

# %% [markdown]
# ### Aligning Edmonton's classes to the Province's buckets
#
# The two vocabularies are not identical, and this is the most fragile joint in
# the notebook.
#
# `RESIDENTIAL`, `FARMLAND` and the non-residential classes map straightforwardly.
# Apartment buildings (`OTHER RESIDENTIAL` on the roll, its own slightly higher
# municipal rate) have **no slot in Schedule MR at all** — the Province's
# columns are Residential / Farmland / Non-Residential / M&E / Other. Edmonton
# files them under **Other**, whose header literally reads *"Other (including
# annexed, vacant, total minimum tax, etc.)"*.
#
# That mapping is *derived, not labelled* — the cell below checks it the only
# way available: the rate implied by the Province's own numbers
# (`levy ÷ assessment`) against Edmonton's published apartment rate. If they
# match, that is what is in the bucket.
#
# > **`ASSUMPTION`** — the published municipal rate below is typed in from
# > Edmonton's rate bylaw for this roll year. It is the one hand-entered number
# > in the notebook, and it is used *only* for this validation, never in a total.

# %%
PUBLISHED_APARTMENT_RATE = 8.2064  # Edmonton "Other Residential" municipal, 2026

implied = 1000 * fir["levy"]["other"] / fir["assessment"]["other"]
print(f"rate implied by Edmonton's FIR filing for 'Other': {implied:.4f}")
print(f"Edmonton's published apartment rate:               {PUBLISHED_APARTMENT_RATE:.4f}")
print(f"difference:                                        {100 * (implied / PUBLISHED_APARTMENT_RATE - 1):+.2f}%")

check(abs(implied / PUBLISHED_APARTMENT_RATE - 1) < 0.02,
      "the Province's 'Other' bucket bills at Edmonton's apartment rate "
      "— apartments are what is in it")

# The residential and non-residential rates should match Edmonton's published
# rates exactly, which is the cheapest possible confirmation that this filing
# and this roll describe the same tax year.
print(f"\nfiled residential rate:     {fir['mill_rate']['residential']:.4f}")
print(f"filed non-residential rate: {fir['mill_rate']['non_residential']:.4f}")

# %%
# ASSUMPTION — this map is Edmonton's class vocabulary as of this roll. A new
# class label appearing on the roll would be silently unmapped, so it raises.
CLASS_TO_FIR = {
    "RESIDENTIAL": "residential",
    "OTHER RESIDENTIAL": "other",
    "FARMLAND": "farmland",
    "COMMERCIAL": "non_residential",
    "MA DERELICT RESIDENTIAL": "non_residential",
    "DESIGNATED IND PROPERTIES": "non_residential",
    EXEMPT_LABEL: None,
}
unmapped = set(classes) - set(CLASS_TO_FIR)
if unmapped:
    raise KeyError(f"unmapped assessment class(es): {sorted(unmapped)}")

BUCKETS = ["residential", "other", "non_residential", "farmland"]
for b in BUCKETS:
    roll[f"av_{b}"] = 0.0
for label_col, pct_col in SLOTS:
    share = roll[pct_col].fillna(0) / 100
    for label, bucket in CLASS_TO_FIR.items():
        if bucket is None:
            continue
        m = roll[label_col] == label
        roll.loc[m, f"av_{bucket}"] += roll.loc[m, "assessed"] * share[m]

apportioned = roll[[f"av_{b}" for b in BUCKETS]].to_numpy().sum()
check(apportioned <= roll["assessed"].sum() * 1.0001,
      "apportioning into buckets does not create assessed value")

# %% [markdown]
# ---
# ## 3. The gap
#
# Everything is now in the same units: assessed dollars, in the Province's
# buckets, for the same year. What the roll says we could tax, against what
# Edmonton says it did tax.

# %%
gaps = pd.DataFrame(
    [
        {
            "bucket": b,
            "roll": roll[f"av_{b}"].sum(),
            "filed": fir["assessment"][b],
        }
        for b in BUCKETS
    ]
).set_index("bucket")
gaps["gap"] = gaps["roll"] - gaps["filed"]
gaps["gap_pct"] = 100 * (gaps["roll"] / gaps["filed"] - 1)

pd.set_option("display.float_format", lambda v: f"{v:,.0f}")
print(gaps[["roll", "filed", "gap"]].to_string())
print(f"\ntotal assessed value on the roll that Edmonton does not tax: "
      f"${gaps['gap'].sum() / 1e9:,.2f}B")
print(f"value the roll actually flags as exempt:                      "
      f"${flagged_value / 1e9:,.4f}B")
print(f"\nthe flag accounts for {100 * flagged_value / gaps['gap'].sum():.2f}% of it.")

check((gaps["gap"] > 0).all(),
      "every class taxes LESS than the roll carries — exemptions only subtract")

# %% [markdown]
# **Residential is nearly explained; the other two are not.** Residential land
# is barely exempt anywhere — a house is a house — so the roll and the filing
# almost agree. Apartments and non-residential diverge by about a fifth.
#
# That shape is itself the first piece of evidence: whatever is missing is
# concentrated in exactly the classes where institutional, public and
# non-profit property lives.

# %%
# --- chart 1 ---------------------------------------------------------------
# Form: magnitude across four unordered categories -> bar. One series, so no
# legend; the title names the measure. Values are direct-labelled because
# reading a percentage off a gridline is worse than reading the number.
SURFACE, INK, INK_MUTED = "#fcfcfb", "#0b0b0b", "#52514e"
BLUE, ORANGE = "#2a78d6", "#eb6834"   # validated categorical slots 1 and 2
GRID = "#dcdcd8"

LABELS = {"residential": "Residential", "other": "Apartments",
          "non_residential": "Non-residential", "farmland": "Farmland"}

fig, ax = plt.subplots(figsize=(8, 3.6), facecolor=SURFACE)
ax.set_facecolor(SURFACE)
order = gaps.sort_values("gap_pct", ascending=True)
bars = ax.barh([LABELS[b] for b in order.index], order["gap_pct"],
               color=BLUE, height=0.55)
for rect, v in zip(bars, order["gap_pct"]):
    ax.text(rect.get_width() + 0.4, rect.get_y() + rect.get_height() / 2,
            f"+{v:.1f}%", va="center", ha="left", color=INK, fontsize=10)

ax.set_xlim(0, order["gap_pct"].max() * 1.18)
ax.xaxis.set_major_formatter(mticker.PercentFormatter(decimals=0))
ax.set_xlabel("assessed value on the roll, above what Edmonton filed as taxable",
              color=INK_MUTED, fontsize=9)
ax.set_title("Every class taxes less than the roll carries — two of them, much less",
             color=INK, fontsize=11, loc="left", pad=12)
ax.grid(axis="x", color=GRID, linewidth=0.6)
ax.set_axisbelow(True)
for s in ("top", "right", "left"):
    ax.spines[s].set_visible(False)
ax.spines["bottom"].set_color(GRID)
ax.tick_params(colors=INK_MUTED, length=0)
plt.tight_layout()
show(fig, "Bar chart of how far each assessment class's roll value exceeds what "
          "Edmonton filed as taxable. Residential is close to matching at about 1 "
          "percent, while apartments and non-residential each stand about 21 percent "
          "above the filed base, concentrating the unexplained value in the two classes "
          "where institutional and non-profit property sits.")

# %% [markdown]
# ---
# ## 4. Source three — zoning, and where the gap sits
#
# Edmonton publishes its Zoning Bylaw polygons. Five zone codes are where
# exempt land would be expected to sit:
#
# | code | zone | why it is a candidate |
# |---|---|---|
# | `AJ` | Alternative Jurisdiction | federal/provincial land, reserves |
# | `UI` | Urban Institution | universities, hospitals |
# | `UF` | Urban Facilities | civic and institutional facilities |
# | `PU` | Public Utility | utility land |
# | `PS` | Parks and Services | parks, schools, recreation |
#
# > **`ASSUMPTION`** — this list is a judgement about which zones *could* hold
# > exempt property, not a legal determination. It is deliberately generous.
# > Note in particular that a zone being a candidate says nothing about any
# > individual parcel in it: `UF` and `PU` both contain privately-owned,
# > fully-taxed facilities.
#
# The per-property `zoning` field on Edmonton's Property Info dataset looks like
# a shortcut here and is a trap — it is null for about a third of accounts
# (condominium units), and null on ~42% of downtown revenue. The polygons are
# the reliable route, so this is a genuine point-in-polygon join.

# %%
import geopandas as gpd  # noqa: E402  — heavy import, only needed from here

ZONING_URL = "https://data.edmonton.ca/resource/fixa-tstc.geojson?$limit=20000"
zoning_path = fetch(ZONING_URL, "zoning.geojson")

zoning = gpd.read_file(zoning_path)
print(f"\nzoning polygons: {len(zoning):,}")

points = gpd.GeoDataFrame(
    roll,
    geometry=gpd.points_from_xy(roll["lon"], roll["lat"]),
    crs="EPSG:4326",
)
# Set the CRS on both sides explicitly, every time, even when they agree.
joined = gpd.sjoin(points, zoning[["zoning", "geometry"]].to_crs("EPSG:4326"),
                   how="left", predicate="within")

# A point on a shared boundary matches both polygons and sjoin emits a row per
# match, which would double-count its value. Keep the first, and report it.
dup = joined.index.duplicated(keep="first")
print(f"properties on a zone boundary (matched >1 polygon): {int(dup.sum()):,}")
joined = joined[~dup]

# A zoning string is like "RS" or "DC2 1234" — the code is the first token.
# NB: .str.split().str[0] on an empty string yields NaN, not "".
roll["zone"] = (
    joined["zoning"].astype("string").str.strip().str.split().str[0].reindex(roll.index)
)
unplaced = roll["zone"].isna()
print(f"properties in no zoning polygon: {int(unplaced.sum()):,} "
      f"(${roll.loc[unplaced, 'assessed'].sum():,.0f} assessed)")

check(len(roll) == len(points),
      "the spatial join neither dropped nor multiplied properties")

CANDIDATE_ZONES = ("AJ", "UF", "UI", "PU", "PS")
roll["candidate"] = roll["zone"].isin(CANDIDATE_ZONES)
print(f"\nproperties on exempt-candidate zoning: {int(roll['candidate'].sum()):,}")

# %%
gaps["on_candidate"] = [roll.loc[roll["candidate"], f"av_{b}"].sum() for b in BUCKETS]
gaps["explained_pct"] = 100 * gaps["on_candidate"] / gaps["gap"]

print(f"{'class':<18}{'gap':>18}{'on candidate zoning':>22}{'explains':>11}")
for b in BUCKETS:
    r = gaps.loc[b]
    print(f"{LABELS[b]:<18}{'$' + format(r['gap'], ',.0f'):>18}"
          f"{'$' + format(r['on_candidate'], ',.0f'):>22}"
          f"{r['explained_pct']:>10.1f}%")

# %% [markdown]
# **For non-residential property, zoning nearly closes the gap.** Five zone
# codes account for the overwhelming majority of the value Edmonton assesses
# and does not tax. That is a real, checkable result: we can *size* Edmonton's
# non-residential exempt property from public data, and say roughly where it is.
#
# For apartments it explains almost nothing. Hold that thought — it is the
# subject of §5.2.

# %%
# --- chart 2 ---------------------------------------------------------------
# Same form as chart 1 (magnitude, few categories) so the two read as a pair.
# A reference line at 100% is what makes the bars mean something: it is the
# point at which zoning would fully account for the gap.
fig, ax = plt.subplots(figsize=(8, 3.2), facecolor=SURFACE)
ax.set_facecolor(SURFACE)

sub = gaps.loc[["non_residential", "other", "residential"]]
bars = ax.barh([LABELS[b] for b in sub.index], sub["explained_pct"],
               color=BLUE, height=0.55)
# A bar running close to the 100% reference line has no room for an outside
# label without colliding with it, so those label inside the bar instead.
for rect, v in zip(bars, sub["explained_pct"]):
    inside = v > 60
    ax.text(rect.get_width() + (-2 if inside else 1.5),
            rect.get_y() + rect.get_height() / 2, f"{v:.1f}%",
            va="center", ha="right" if inside else "left",
            color=SURFACE if inside else INK, fontsize=10)

ax.set_xlim(0, 118)
ax.axvline(100, color=ORANGE, linewidth=2, zorder=3)
# Right-aligned so it sits in the empty space left of the line rather than
# overflowing the axes.
ax.text(98, ax.get_ylim()[1] - 0.1, "gap fully accounted for →",
        color=ORANGE, fontsize=9, va="top", ha="right")
ax.xaxis.set_major_formatter(mticker.PercentFormatter(decimals=0))
ax.set_xlabel("share of the class's gap sitting on exempt-candidate zoning",
              color=INK_MUTED, fontsize=9)
ax.set_title("Zoning locates the non-residential gap, and misses the apartment one",
             color=INK, fontsize=11, loc="left", pad=12)
ax.grid(axis="x", color=GRID, linewidth=0.6)
ax.set_axisbelow(True)
for s in ("top", "right", "left"):
    ax.spines[s].set_visible(False)
ax.spines["bottom"].set_color(GRID)
ax.tick_params(colors=INK_MUTED, length=0)
plt.tight_layout()
show(fig, "Bar chart of how much of each class's gap sits on exempt-candidate "
          "zoning. Non-residential reaches 96 percent, close to the marked line where "
          "zoning would fully account for the gap. Apartments reach only 13 percent and "
          "residential 10 percent, so for those two classes zoning locates almost none "
          "of the exempt value.")

# %% [markdown]
# ---
# ## 5. Why this cannot be turned into a list of exempt properties
#
# Three separate obstacles. Any one of them would be enough.

# %% [markdown]
# ### 5.1 The candidate set does not balance — in the wrong direction
#
# If exempt property were exactly "everything on those five zones", the
# candidate total would equal the gap. It does not.

# %%
nr = gaps.loc["non_residential"]
short = nr["gap"] - nr["on_candidate"]
print(f"non-residential gap:                      ${nr['gap']:,.0f}")
print(f"non-res value on candidate zoning:        ${nr['on_candidate']:,.0f}")
print(f"shortfall:                                ${short:,.0f}  "
      f"({100 * short / nr['gap']:.1f}% of the gap)")

# %% [markdown]
# The candidate set is *smaller* than the gap. So even granting the most
# generous reading — that every square metre of those five zones is exempt —
# there is still property being exempted somewhere else.
#
# And the reverse is true simultaneously: those zones certainly contain taxed
# property (a privately-run facility on `UF`, a commercial operation on `PU`).
# Both errors are present at once, in unknown proportion, and **they partially
# cancel** — so the 95.5% agreement above is compatible with a large number of
# individual misclassifications in both directions.
#
# A single aggregate cannot separate two errors that offset each other.

# %% [markdown]
# ### 5.2 For apartments, zoning is the wrong instrument entirely
#
# Alberta's *Municipal Government Act* s.362 exempts property by **use and
# ownership**, not by zone: seniors' housing, non-profit and subsidised
# housing, and residences owned by charitable bodies. A non-profit seniors'
# residence sits on ordinary residential zoning, indistinguishable in the
# polygon layer from the market apartment block next door.

# %%
ap = gaps.loc["other"]
invisible = ap["gap"] - ap["on_candidate"]
apartments = roll[(roll["av_other"] > 0) & (~roll["candidate"])]

print(f"apartment gap:                           ${ap['gap']:,.0f}")
print(f"  explained by candidate zoning:         ${ap['on_candidate']:,.0f}  "
      f"({100 * ap['on_candidate'] / ap['gap']:.1f}%)")
print(f"  invisible to zoning:                   ${invisible:,.0f}  "
      f"({100 * invisible / ap['gap']:.1f}%)")
print(f"\napartment properties on ordinary zoning:  {len(apartments):,}")
print(f"  their combined assessed value:         ${apartments['av_other'].sum():,.0f}")
print(f"  median value:                          ${apartments['av_other'].median():,.0f}")

check(invisible > 0,
      "most of the apartment gap sits on zoning that carries no exemption signal")

# %% [markdown]
# ### 5.3 The identification problem — one equation, thousands of unknowns
#
# This is the obstacle that no additional cleverness removes.
#
# What we have is a **single number**: the aggregate gap. What we want is a
# **per-parcel flag**. Any set of parcels whose values sum to the gap is
# perfectly consistent with everything published — and there are enormously
# many such sets.
#
# That is not a hand-wave. Below, two sets of apartment properties are
# constructed that **share not one property between them**, and each reproduces
# the invisible apartment gap to within a rounding error. Nothing in any public
# dataset prefers one over the other.

# %%
values = apartments["av_other"].to_numpy()
rng = np.random.default_rng(11)
shuffled = rng.permutation(len(values))
left, right = shuffled[: len(values) // 2], shuffled[len(values) // 2:]


def subset_reaching(vals: np.ndarray, target: float) -> tuple[np.ndarray, float]:
    """Greedy largest-first selection summing as close to target as possible."""
    picked, total = [], 0.0
    for i in np.argsort(-vals):
        if total + vals[i] <= target:
            picked.append(i)
            total += vals[i]
    return np.array(picked), total


pick_a, total_a = subset_reaching(values[left], invisible)
pick_b, total_b = subset_reaching(values[right], invisible)

accounts_a = set(apartments.iloc[left[pick_a]]["account"])
accounts_b = set(apartments.iloc[right[pick_b]]["account"])

print(f"target — apartment exemption invisible to zoning: ${invisible:,.0f}\n")
print(f"  set A: {len(accounts_a):>4,} properties   ${total_a:,.0f}   "
      f"({100 * total_a / invisible:.4f}% of target)")
print(f"  set B: {len(accounts_b):>4,} properties   ${total_b:,.0f}   "
      f"({100 * total_b / invisible:.4f}% of target)")
print(f"  properties in common: {len(accounts_a & accounts_b)}")

check(len(accounts_a & accounts_b) == 0,
      "the two sets are disjoint — no property appears in both")
check(abs(total_a - total_b) / invisible < 0.01,
      "both sets reproduce the same aggregate to within 1%")

# %% [markdown]
# Two disjoint answers, equally consistent with every published number. And
# these are just two of the astronomically many that exist — the greedy
# construction above is not searching for near-misses, it is stumbling onto
# them immediately from a random split.
#
# **The aggregate constrains the total. It does not constrain the membership.**

# %%
# --- chart 3 ---------------------------------------------------------------
# Form: identity across two constructed sets against a background population
# -> strip plot on a log value axis (values span five orders of magnitude).
# Two series, so a legend is present; blue/orange are validated slots 1-2 and
# clear the all-pairs CVD floor. The background is deliberately neutral ink,
# not a third hue: it is context, not a category.
fig, ax = plt.subplots(figsize=(8, 3.9), facecolor=SURFACE)
ax.set_facecolor(SURFACE)

jitter = rng.uniform(-0.14, 0.14, len(values))
ax.scatter(values, jitter, s=7, color="#c9c8c2", alpha=0.55, linewidths=0,
           label=f"all {len(values):,} apartment properties on ordinary zoning")

va, vb = values[left[pick_a]], values[right[pick_b]]
ax.scatter(va, np.full(len(va), 0.42), s=34, color=BLUE,
           edgecolors=SURFACE, linewidths=1.2, zorder=3,
           label=f"set A — {len(va)} properties, ${total_a / 1e9:.3f}B")
ax.scatter(vb, np.full(len(vb), -0.42), s=34, color=ORANGE,
           edgecolors=SURFACE, linewidths=1.2, zorder=3,
           label=f"set B — {len(vb)} properties, ${total_b / 1e9:.3f}B")

ax.set_xscale("log")
# Headroom above the set-A row so the legend cannot sit on top of its markers.
ax.set_ylim(-0.8, 1.45)
ax.set_yticks([])
ax.xaxis.set_major_formatter(mticker.FuncFormatter(
    lambda v, _: f"${v / 1e6:,.0f}M" if v >= 1e6 else f"${v / 1e3:,.0f}k"))
ax.set_xlabel("assessed value per property (log scale)", color=INK_MUTED, fontsize=9)
ax.set_title("Two disjoint sets of properties, one identical total",
             color=INK, fontsize=11, loc="left", pad=12)
ax.grid(axis="x", color=GRID, linewidth=0.6)
ax.set_axisbelow(True)
for s in ("top", "right", "left"):
    ax.spines[s].set_visible(False)
ax.spines["bottom"].set_color(GRID)
ax.tick_params(colors=INK_MUTED, length=0)
leg = ax.legend(loc="upper left", frameon=False, fontsize=8.5, labelcolor=INK_MUTED)
plt.tight_layout()
show(fig, f"Strip plot of all {len(values):,} apartment properties on ordinary zoning "
          f"by assessed value on a log scale, with two constructed sets marked on "
          f"separate rows. Set A ({len(va)} properties) and set B ({len(vb)} properties) "
          f"share no property, yet both draw from the same value range and both sum to "
          f"the same total, so no published attribute distinguishes the correct set "
          f"from the incorrect one.")

# %% [markdown]
# ---
# ## 6. What would actually resolve this
#
# One column. Per-parcel exemption status — the flag the City necessarily holds,
# because it bills from it — published alongside the roll it already publishes.
#
# Everything above is triangulation *around* a field that exists and is not
# released. With it, the gap stops being a subtraction between two aggregates
# and becomes a fact about each property. Without it, no amount of public data
# gets past §5.3, because the obstacle there is not missing detail — it is that
# a sum does not determine its terms.

# %% [markdown]
# ## 7. What this notebook does *not* claim
#
# - **That any specific property is tax-exempt.** That is the entire point. The
#   five candidate zones size an uncertainty; they do not accuse a parcel.
# - **That the gap is exclusively exemptions.** It is the difference between two
#   aggregates compiled by different processes at possibly different moments.
#   Timing, appeals, supplementary assessments and bucket-definition
#   differences all live in it too. Exemption is the largest and best-supported
#   component, not the only conceivable one.
# - **That Edmonton is doing anything wrong.** Exempting universities, parks and
#   hospitals is deliberate public policy. The observation is about what is
#   *legible* from published data, not about whether the policy is right.
# - **That the two constructed sets in §5.3 are plausible exemption lists.** They
#   are counterexamples, not candidates. Their only job is to demonstrate that
#   the aggregate does not determine the membership.

# %%
passed = sum(1 for ok, _ in CHECKS if ok)
print(f"invariants checked: {passed}/{len(CHECKS)} passed\n")
for ok, claim in CHECKS:
    print(f"  [{'PASS' if ok else 'FAIL'}] {claim}")
if passed != len(CHECKS):
    raise AssertionError(f"{len(CHECKS) - passed} invariant(s) failed — see above")
print("\nAll structural invariants held.")

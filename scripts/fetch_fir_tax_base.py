"""Fetch Edmonton's filed taxable assessment + municipal levy from Alberta FIR.

The province publishes, per municipality-year, what the City actually filed:
Schedule MR's taxable assessment (``MR(2)``), municipal tax levy (``MR(1)``)
and mill rates (``MR(3)``). Same workbooks ``fetch_fir_debt.py`` already pulls
for Schedule AA — that script reads one sheet of 51; this reads three more.

    https://open.alberta.ca/opendata/municipal-financial-and-statistical-data

Why this exists (2026-08-25): the assessment roll's year is NOT reliably
recoverable from Socrata metadata. Edmonton left ``Period of Coverage`` reading
2025 for the whole 2026 roll, so ``check_year_alignment.py`` — which parsed
that string — reported "aligned" while the pipeline billed a 2026 roll at 2025
rates. Residential taxable assessment is the fix: residential land is barely
exempt anywhere, so our residential base should track the filed one within a
couple of percent, and it is a year apart from its neighbours by ~10%. That
makes it a sharp year detector. ``check_roll_year_against_fir.py`` consumes
this file offline.

⚠️ MANUAL, REVIEWED INPUT (the mill-rates / fir_debt_series pattern) — NOT part
of the weekly refresh. Re-run when a new financial year publishes (~annually),
eyeball the diff, commit. ``openpyxl`` is a dev-only dep.

⚠️ The buckets are NOT 1:1 with our tax classes. ``MR(2)`` column [10] is
"Other (including annexed, vacant, total minimum tax, etc.)", but Edmonton has
no apartment slot in MR and files its Other Residential sub-class there — the
implied rate (levy / assessment) lands within 1% of our Other Residential rate,
which is how that was established. Only the residential column is safe to read
directly, and it is the only one the year check uses.

Usage:
    .venv/bin/python scripts/fetch_fir_tax_base.py [--out data/fir_tax_base.json]
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import tempfile
from datetime import date
from pathlib import Path

import requests

logger = logging.getLogger(__name__)

DATASET_PAGE = (
    "https://open.alberta.ca/opendata/municipal-financial-and-statistical-data"
)
OUT_PATH = Path(__file__).resolve().parent.parent / "data" / "fir_tax_base.json"

EDMONTON_CODE = "0098"
EDMONTON_NAME = "EDMONTON"

# Column index -> field name within Schedule MR's municipality rows. Positions
# are stable across the yearly workbooks (header row 1 names them); the loader
# cross-checks the header text and raises on drift rather than reading blind.
MR_COLUMNS = {
    5: "residential",
    6: "farmland",
    7: "non_residential",
    9: "machinery_equipment",
    10: "other",
}

# Substring each column's header must contain, lowercased. Guards against the
# province reordering Schedule MR — a silent column shift would corrupt the one
# number the year detector trusts.
MR_HEADER_EXPECT = {
    5: "residential",
    6: "farmland",
    7: "non-residential",
    9: "machinery",
    10: "other",
}

SHEETS = {"MR(1)-Tax Levy": "levy", "MR(2)-Assessment": "assessment",
          "MR(3)-Mill Rate": "mill_rate"}


def discover_resources(page_html: str) -> dict[int, str]:
    """Map financial year -> workbook URL from the dataset page.

    Also picks up ``YYYY_tax_rates.xlsx``, which is how the province publishes
    the newest year's Schedule MR before its full financial workbook lands.
    """
    urls = set(re.findall(r'href="(https://open\.alberta\.ca/dataset/[^"]+)"', page_html))
    out: dict[int, str] = {}
    for u in urls:
        m = re.search(r"/download/(\d{4})_(?:financial_year|tax_rates)\.xlsx$", u, re.I)
        if m:
            # A financial_year workbook wins over a tax_rates one for the same
            # year: it is the complete filing, not the early rate release.
            year = int(m.group(1))
            if year not in out or "financial_year" in u.lower():
                out[year] = u
    return out


def _edmonton_row(ws) -> tuple:
    for row in ws.iter_rows(values_only=True):
        if any(isinstance(c, str) and c.strip().upper() == EDMONTON_NAME for c in row[:6]):
            if str(row[2]).strip() != EDMONTON_CODE:
                raise ValueError(
                    f"{EDMONTON_NAME} found under code {row[2]!r}, expected {EDMONTON_CODE}"
                )
            return row
    raise ValueError(f"no {EDMONTON_NAME} row in {ws.title!r}")


def _check_header(ws) -> None:
    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    for idx, expect in MR_HEADER_EXPECT.items():
        got = str(header[idx] or "").lower()
        if expect not in got:
            raise ValueError(
                f"{ws.title!r} column {idx} header is {header[idx]!r}, "
                f"expected to contain {expect!r} — Schedule MR moved, do not read blind"
            )


def parse_workbook(path: Path) -> dict:
    import openpyxl  # noqa: PLC0415 — dev-only dep, imported at use

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out: dict[str, dict[str, float]] = {}
        for sheet, key in SHEETS.items():
            if sheet not in wb.sheetnames:
                raise ValueError(f"{path.name}: no {sheet!r} sheet")
            ws = wb[sheet]
            _check_header(ws)
            row = _edmonton_row(ws)
            out[key] = {
                name: (None if row[idx] is None else float(row[idx]))
                for idx, name in MR_COLUMNS.items()
            }
        return out
    finally:
        wb.close()


def check_internal_consistency(rec: dict, year: int) -> None:
    """assessment x mill_rate must reproduce levy — proves MR(2) is the TAXABLE base.

    This is the check that makes the file trustworthy as an anchor: if the
    province's own three sheets agree to a rounding error, the assessment
    column is what the levy was actually computed on, not a gross roll total.
    """
    for field in ("residential", "non_residential"):
        a = rec["assessment"].get(field)
        r = rec["mill_rate"].get(field)
        lv = rec["levy"].get(field)
        if not a or not r or lv is None:
            continue
        calc = a * r / 1000
        if lv and abs(calc - lv) / lv > 1e-4:
            raise ValueError(
                f"{year} {field}: assessment x rate = {calc:,.0f} but levy = {lv:,.0f} "
                f"({100 * (calc / lv - 1):+.4f}%) — MR sheets disagree"
            )


def main(argv=None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--out", type=Path, default=OUT_PATH)
    p.add_argument("--first-year", type=int, default=2023,
                   help="earliest workbook to pull (default: 2023)")
    args = p.parse_args(argv)

    page = requests.get(DATASET_PAGE, timeout=120)
    page.raise_for_status()
    res = discover_resources(page.text)
    years = sorted(y for y in res if y >= args.first_year)
    if not years:
        raise SystemExit(f"no workbooks at/after {args.first_year} on {DATASET_PAGE}")
    logger.info("found workbooks for %s", years)

    out = {
        "_source": DATASET_PAGE,
        "_licence": "Open Government Licence – Alberta",
        "_municipality": {"name": EDMONTON_NAME, "fir_code": EDMONTON_CODE},
        "_fetched": date.today().isoformat(),
        "_note": (
            "Alberta FIR Schedule MR — what Edmonton FILED with the province. "
            "MR(2) is the TAXABLE base (assessment x MR(3) reproduces MR(1); "
            "asserted by check_internal_consistency at fetch time). Only "
            "`residential` is safe to compare 1:1 against our tax classes."
        ),
        "years": {},
    }
    with tempfile.TemporaryDirectory() as tmp:
        for year in years:
            dest = Path(tmp) / f"{year}.xlsx"
            logger.info("downloading %s", res[year])
            r = requests.get(res[year], timeout=600)
            r.raise_for_status()
            dest.write_bytes(r.content)
            rec = parse_workbook(dest)
            check_internal_consistency(rec, year)
            out["years"][str(year)] = rec
            logger.info("  %d residential taxable base $%s", year,
                        f"{rec['assessment']['residential']:,.0f}")

    args.out.write_text(json.dumps(out, indent=2) + "\n")
    logger.info("wrote %s (%d years)", args.out, len(out["years"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

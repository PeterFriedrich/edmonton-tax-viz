"""Fetch the Alberta FIR long-term debt series for Edmonton + peer municipalities.

Debt-lens ticket D5 (docs/fable_brief_debt_lens.md): the citywide debt context
annotation needs Edmonton's long-term debt time series plus the two
already-tracked peers (St. Albert, Strathcona County) for benchmarking. The
source is the province's Municipal Financial and Statistical Data (FIR/SIR):

    https://open.alberta.ca/opendata/municipal-financial-and-statistical-data

one XLSX workbook per financial year, every Alberta municipality. The debt
schedule ("Schedule AA" / "GR Debt Info" in 2003) carries four fields per
municipality-year, with FIR item codes:

    Debt Limit               05700   (MGA Debt Limit Regulation 255/2000 —
                                      NOT the City's own DMFP policy limits)
    Total Debt               05710
    Debt Service Limit       05720
    Total Debt Service Costs 05730

Format eras (all verified 2026-07-14):
  - 2009+  : one ``YYYY_financial_year.xlsx`` workbook, sheet ``AA(1)-Debt``
             (2009-2016 arrive inside one zip).
  - 2004-08: per-schedule files inside the 2003-2008 zip,
             ``YYYY/YYYY-AA-Debt Info.xlsx``, sheet ``Schedule AA``.
  - 2003   : legacy ``.xls``, ``2003-EA-MR/GR Debt Info.xls`` (needs xlrd).
Header row and column names are IDENTICAL across all eras.

This is a **manual, reviewed input** (mill-rates pattern — see DATA.md §11):
run it by hand when a new financial year publishes (~annually), eyeball the
diff on ``data/fir_debt_series.json``, commit. It is NOT part of the weekly
refresh workflow, and openpyxl/xlrd are dev-only deps (requirements.txt, not
requirements-ci.txt).

Integrity rules (no silent data drops):
  - a missing year, municipality row, or column HARD-FAILS;
  - every value is sanity-checked against its year-neighbours (factor-5 band)
    to catch source-side unit slips — the known one is Strathcona County 2013,
    reported in $000s (KNOWN_UNIT_CORRECTIONS, applied and recorded in the
    output's notes);
  - two anchor values cross-check the series against independently published
    figures; drift means the province restated data and a human must look
    (--allow-anchor-drift to accept after review).

Usage:
    python scripts/fetch_fir_debt.py                    # download + extract
    python scripts/fetch_fir_debt.py --cache-dir /tmp/fir   # reuse downloads
"""

import argparse
import json
import logging
import re
import sys
import zipfile
from datetime import date
from pathlib import Path

import requests

logger = logging.getLogger(__name__)

DATASET_PAGE = (
    "https://open.alberta.ca/opendata/municipal-financial-and-statistical-data"
)
OUT_PATH = Path(__file__).resolve().parent.parent / "data" / "fir_debt_series.json"

FIRST_YEAR = 2003  # brief's horizon; older zips exist (1994+) if ever wanted

# Extracted by stable municipal CODE; the name is cross-checked and a mismatch
# warns (names could drift in the source, codes should not).
MUNICIPALITIES = {
    "0098": "EDMONTON",
    "0292": "ST. ALBERT",
    "0302": "STRATHCONA COUNTY",
}

# Output field name -> exact source column header (identical 2003-2025).
FIELDS = {
    "debt_limit": "Debt Limit",
    "total_debt": "Total Debt",
    "debt_service_limit": "Debt Service Limit",
    "total_debt_service_costs": "Total Debt Service Costs",
}

# (code, year) -> multiplier. Strathcona County's 2013 row is reported in
# $000s in the source workbook (debt limit 485,926 sits between real-dollar
# neighbours 473.9M and 504.2M) — a source-side unit slip, corrected here and
# recorded in the output notes.
KNOWN_UNIT_CORRECTIONS = {("0302", 2013): 1000}

# Independently published figures the extracted series must reproduce:
#   - Edmonton 2025 total debt: "$4.6 billion" reported to Council 2026-03-17
#     (docs/fable_brief_debt_lens.md, Component 2 headline);
#   - Strathcona County 2022 total debt: $133.07M, 2022 audited statements
#     (the brief's peer datapoint).
ANCHORS = {
    ("0098", 2025, "total_debt"): 4_592_150_000,
    ("0302", 2022, "total_debt"): 133_070_148,
}

NEIGHBOUR_BAND = 5.0  # value vs adjacent years must stay within this factor


def discover_resources(page_html):
    """Map year -> download URL from the dataset page; zips keyed by era."""
    urls = set(re.findall(r'href="(https://open\.alberta\.ca/dataset/[^"]+)"', page_html))
    out = {"years": {}, "zips": {}}
    for u in urls:
        m = re.search(r"/download/(\d{4})_financial_year\.xlsx$", u, re.I)
        if m:
            out["years"][int(m.group(1))] = u
        elif u.lower().endswith("2009-2016-municipal_financial-data-and-statistics.zip"):
            out["zips"]["2009-2016"] = u
        elif u.lower().endswith("xlsx-2003-2008.zip"):
            out["zips"]["2003-2008"] = u
    return out


def download(url, dest):
    if dest.exists() and dest.stat().st_size > 0:
        logger.info("cached: %s", dest.name)
        return dest
    logger.info("downloading %s", url)
    r = requests.get(url, timeout=300)
    r.raise_for_status()
    dest.write_bytes(r.content)
    return dest


def _extract_rows(header_and_rows, year, source_name):
    """Pull the wanted municipality rows out of one debt schedule.

    ``header_and_rows``: iterable of row tuples (raw cell values). Returns
    {code: {field: int}} and hard-fails on any missing piece.
    """
    header = None
    col = {}
    found = {}
    for row in header_and_rows:
        vals = ["" if c is None else str(c).strip() for c in row]
        if header is None:
            if "MUNICIPALITY" in vals:
                header = vals
                for out_name, src_name in FIELDS.items():
                    if src_name not in header:
                        raise SystemExit(
                            f"{source_name}: column {src_name!r} missing from header {header}"
                        )
                    col[out_name] = header.index(src_name)
                col["_code"] = header.index("CODE")
                col["_name"] = header.index("MUNICIPALITY")
            continue
        code = vals[col["_code"]]
        if code.endswith(".0"):  # numeric cell in the legacy .xls era
            code = code[:-2]
        code = code.zfill(4) if code else ""
        if code in MUNICIPALITIES:
            name = vals[col["_name"]].upper()
            if name != MUNICIPALITIES[code]:
                logger.warning(
                    "%s: code %s named %r (expected %r) — check the source",
                    source_name, code, name, MUNICIPALITIES[code],
                )
            rec = {}
            for out_name in FIELDS:
                raw = vals[col[out_name]]
                try:
                    rec[out_name] = int(round(float(raw)))
                except ValueError:
                    raise SystemExit(
                        f"{source_name}: non-numeric {out_name}={raw!r} for {name} {year}"
                    )
            mult = KNOWN_UNIT_CORRECTIONS.get((code, year))
            if mult:
                rec = {k: v * mult for k, v in rec.items()}
                rec["unit_corrected"] = mult
                logger.warning(
                    "%s: applied documented x%d unit correction for %s %s",
                    source_name, mult, MUNICIPALITIES[code], year,
                )
            found[code] = rec
    if header is None:
        raise SystemExit(f"{source_name}: no MUNICIPALITY header row found")
    missing = set(MUNICIPALITIES) - set(found)
    if missing:
        raise SystemExit(
            f"{source_name}: municipalities missing for {year}: "
            f"{[MUNICIPALITIES[c] for c in sorted(missing)]}"
        )
    return found


def parse_xlsx_workbook(path, year):
    """Yearly workbook (2009+): sheet AA(1)-Debt. Also fits Schedule AA files."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        names = [s for s in wb.sheetnames if "debt" in s.lower() or s == "Schedule AA"]
        if not names:
            raise SystemExit(f"{path.name}: no debt sheet in {wb.sheetnames}")
        ws = wb[names[0]]
        return _extract_rows(ws.iter_rows(values_only=True), year, path.name)
    finally:
        wb.close()


def parse_xls_2003(path):
    import xlrd

    book = xlrd.open_workbook(path)
    sh = book.sheet_by_index(0)
    rows = (tuple(sh.cell_value(i, j) for j in range(sh.ncols)) for i in range(sh.nrows))
    return _extract_rows(rows, 2003, path.name)


def sanity_check_neighbours(series):
    """Each value must sit within NEIGHBOUR_BAND of at least one adjacent year.

    Catches unit slips like the (corrected) Strathcona 2013 row if a new one
    ever appears. Zero values are skipped (a real paid-off-debt year would be
    legitimate); anything else fails hard.
    """
    problems = []
    for code, name in MUNICIPALITIES.items():
        years = sorted(series[name]["years"])
        for field in FIELDS:
            vals = {y: series[name]["years"][y][field] for y in years}
            for i, y in enumerate(years):
                v = vals[y]
                if v == 0:
                    continue
                neighbours = [vals[years[j]] for j in (i - 1, i + 1)
                              if 0 <= j < len(years) and vals[years[j]] > 0]
                if not neighbours:
                    continue
                if all(not (1 / NEIGHBOUR_BAND <= v / n <= NEIGHBOUR_BAND)
                       for n in neighbours):
                    problems.append(f"{name} {y} {field}={v:,} vs neighbours "
                                    f"{[f'{n:,}' for n in neighbours]}")
    if problems:
        raise SystemExit(
            "neighbour sanity check failed (possible source unit slip — if real, "
            "add a KNOWN_UNIT_CORRECTIONS entry or adjust NEIGHBOUR_BAND):\n  "
            + "\n  ".join(problems)
        )


def check_anchors(series, allow_drift):
    code_to_name = MUNICIPALITIES
    bad = []
    for (code, year, field), expected in ANCHORS.items():
        name = code_to_name[code]
        got = series[name]["years"].get(year, {}).get(field)
        if got != expected:
            bad.append(f"{name} {year} {field}: got {got:,} expected {expected:,}")
    if bad:
        msg = ("anchor cross-check failed — the province may have restated "
               "figures; review before accepting:\n  " + "\n  ".join(bad))
        if allow_drift:
            logger.warning("%s\n(accepted via --allow-anchor-drift)", msg)
        else:
            raise SystemExit(msg + "\n(re-run with --allow-anchor-drift to accept)")


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--cache-dir", type=Path, default=Path("/tmp/fir"),
                    help="where downloads land / are reused (default /tmp/fir)")
    ap.add_argument("--out", type=Path, default=OUT_PATH)
    ap.add_argument("--allow-anchor-drift", action="store_true",
                    help="accept anchor mismatches after human review")
    args = ap.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    cache = args.cache_dir
    cache.mkdir(parents=True, exist_ok=True)

    page = requests.get(DATASET_PAGE, timeout=120)
    page.raise_for_status()
    res = discover_resources(page.text)
    if not res["years"] or set(res["zips"]) != {"2009-2016", "2003-2008"}:
        raise SystemExit(f"dataset page changed shape: found years "
                         f"{sorted(res['years'])} zips {sorted(res['zips'])}")
    last_year = max(res["years"])
    logger.info("per-year files %s-%s + 2 era zips", min(res["years"]), last_year)

    # --- gather every year's debt table -----------------------------------
    per_year = {}

    for era, url in res["zips"].items():
        zpath = download(url, cache / f"{era}.zip")
        with zipfile.ZipFile(zpath) as z:
            z.extractall(cache / era)

    # 2003: legacy .xls
    xls = next((cache / "2003-2008").rglob("GR Debt Info.xls"), None)
    if xls is None:
        raise SystemExit("2003 'GR Debt Info.xls' not found in the 2003-2008 zip")
    per_year[2003] = parse_xls_2003(xls)

    # 2004-2008: per-schedule workbooks
    for year in range(2004, 2009):
        f = next((cache / "2003-2008").rglob(f"{year}-AA-Debt Info.xlsx"), None)
        if f is None:
            raise SystemExit(f"{year}-AA-Debt Info.xlsx not found in the 2003-2008 zip")
        per_year[year] = parse_xlsx_workbook(f, year)

    # 2009-2016: workbooks inside the era zip
    for year in range(2009, 2017):
        f = next((cache / "2009-2016").rglob(f"{year}_Financial_Year.xlsx"), None)
        if f is None:
            raise SystemExit(f"{year}_Financial_Year.xlsx not found in the 2009-2016 zip")
        per_year[year] = parse_xlsx_workbook(f, year)

    # 2017+: standalone per-year files
    for year in range(2017, last_year + 1):
        if year not in res["years"]:
            raise SystemExit(f"dataset page has no {year}_financial_year.xlsx")
        f = download(res["years"][year], cache / f"{year}.xlsx")
        per_year[year] = parse_xlsx_workbook(f, year)

    missing_years = set(range(FIRST_YEAR, last_year + 1)) - set(per_year)
    if missing_years:
        raise SystemExit(f"years missing from the series: {sorted(missing_years)}")

    # --- reshape to per-municipality series + integrity checks -------------
    series = {}
    for code, name in MUNICIPALITIES.items():
        series[name] = {
            "code": code,
            "years": {y: per_year[y][code] for y in sorted(per_year)},
        }
    sanity_check_neighbours(series)
    check_anchors(series, args.allow_anchor_drift)

    out = {
        "source": {
            "dataset": DATASET_PAGE,
            "schedule": "Schedule AA — Debt Limit / Debt Info "
                        "(item codes 05700/05710/05720/05730); "
                        "2003 from 'GR Debt Info.xls'",
            "retrieved": date.today().isoformat(),
            "years": f"{FIRST_YEAR}-{last_year}",
            "units": "dollars (as reported; see notes for corrections)",
        },
        "notes": [
            "Debt Limit / Debt Service Limit here are the MGA regulation "
            "limits (Debt Limit Regulation 255/2000), NOT Edmonton's own "
            "DMFP policy limits (18%/21% servicing) — do not conflate.",
            "STRATHCONA COUNTY 2013 is reported in $000s in the source "
            "workbook; a documented x1000 correction is applied "
            "(unit_corrected: 1000 on that year's record).",
            "Manual, reviewed input (mill-rates pattern): re-run "
            "scripts/fetch_fir_debt.py when a new financial year publishes, "
            "review the diff, commit.",
        ],
        "municipalities": series,
    }
    args.out.write_text(json.dumps(out, indent=1) + "\n")
    ed = series["EDMONTON"]["years"]
    logger.info(
        "wrote %s: %d years x %d municipalities; EDMONTON total debt %s: $%s -> %s: $%s",
        args.out, len(per_year), len(series),
        FIRST_YEAR, f"{ed[FIRST_YEAR]['total_debt']:,}",
        last_year, f"{ed[last_year]['total_debt']:,}",
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
